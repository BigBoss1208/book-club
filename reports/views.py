from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from members.models import ClubMember, PointLog, Club
from datetime import datetime

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_training_points_view(request):
    now = datetime.now()
    filename = f"diem_ren_luyen_thang_{now.month}_{now.year}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách thành viên"

    headers = ["STT", "Họ tên", "Mã SV", "CLB", "Tổng điểm", "Số buổi tham gia", "Xếp loại"]
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    align = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = align

    members = ClubMember.objects.select_related('user', 'club').all()
    for idx, m in enumerate(members, 1):
        name = m.user.get_full_name() or m.user.username
        student_code = m.user.username
        club = m.club.name if m.club else ""
        total_points = m.total_points
        num_sessions = PointLog.objects.filter(member=m, reason='CHECKIN').count()
        if total_points >= 80:
            rank = "Xuất sắc"
        elif total_points >= 50:
            rank = "Tốt"
        else:
            rank = "Trung bình"
        row = [idx, name, student_code, club, total_points, num_sessions, rank]
        ws.append(row)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=idx+1, column=col)
            cell.border = border
            cell.alignment = align

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
